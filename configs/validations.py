from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter
import os
import shutil
import sys
import json
import datetime
from datetime import datetime as dt
import time
import pandas as pd
import requests
from configs.logging_config import logger


def replace_string(date:datetime, extension, access_path=False, code=0,
                   report_path=False, report_type=None):
    """Informa o final do arquivo (replace removido na verificação se o diretório existe)

    Args:
        - date (datetime): data
        - extension (str): extensão do path para o arquivo desejado
        - access_path (bool, optional): especifica se é para um path do diretório de access. Defaults to False.
        - code (int, optional): código do cliente. Defaults to 0.

    Returns:
        - str: trecho final do path que vai ser removido na verificaçao se o diretório existe
    """
    if access_path:
        return r'token-{}.{}'.format(code, extension)
    
    if report_path:
        return r'{}-report.{}'.format(report_type, extension)

    return r'relatorio-getnet-{}.{}'.format(date.strftime("%B"), extension)

def isUnix(path, **kwargs):
    """Gera um path modificado para cada s.o, aumentando a compatibilidade

    Returns:
      - {str}: path
    """
    refactored_path = path.replace('\\', '/') if (os.name == "posix") else path
    refactored_path += refactored_path.join([f"{value}" for value in kwargs.items()])
    return refactored_path

def path_exists(path_and_replace):
    """Essa classe verifica se o diretório já existe e o cria caso não exista.
    Recebe tanto uma tupla (retono de getPath()), quanto um path direto

    Returns:
      - {str}: path completo
    """
    path = ""
    replace = ""
    if isinstance(path_and_replace, tuple):
        path = isUnix(path_and_replace[0])
        replace = isUnix(path_and_replace[1])

        if not os.path.exists(path):
                os.makedirs(path)

        return path+replace
    
    path = isUnix(path_and_replace)
    if not os.path.exists(path):
        os.makedirs(path)
    
    return path

def getPath(code:int, date:datetime.datetime, extension:str, 
            complete=False, access_path=False, create_if_not_exists=False,
            report_path=False, company:str=None, report_type:str=None, retro=False) -> str:
        """
        Gera o path_file do arquivo solicitado, e retorna uma tupla ou uma string 
        dependendo do parâmetro complete. Caso o complete seja False, retornará uma tupla (path, replace)

        Args:
           - code: (int) código da empresa
           - date: (datetime) data inicial solicitada
           - extension: (str) extensao do arquivo
           - complete: (boolean, optional) certifica o retorno do método como descrito
           - access_path: (boolean, optional) serve para polimorfismo da função, garantindo compatibilidade a quem chama o método
            - create_if_not_exists: (boolean, optional) se informado como True, verifica o diretório e o cria
        
        Returns:
            Sempre string, porém em 2 tipos distintos
            - { tuple } -> (path, replace)
            - { str } -> path
        """
        path = os.getcwd()
        month = date.strftime("%B")
        if not access_path and not report_path:
            path += r'\connection\files\{}\{}\relatorio-getnet-{}.{}'.format(code,
                                                                        date.year,
                                                                        date.strftime("%B"),
                                                                        extension)
            replace = replace_string(date, extension)
            path = path.replace(replace, "")

        if access_path:
            path += r'\connection\access\company-tokens\token-{}.{}'.format(code, extension)
            replace = replace_string(date, extension, access_path=True, code=code)
            path = path.replace(replace, "")
        if report_path and retro:
            path += r'\relatorio\reports\{}\{}\{}-report.{}'.format(company, date.year, report_type, extension)
            replace = replace_string(date, extension, report_path=True, report_type=report_type)
            path = path.replace(replace, "")
        else:
            path += r'\relatorio\reports\{}\{}\{}-{}\{}-report.{}'.format(company, date.year, date.month, month, report_type, extension)
            replace = replace_string(date, extension, report_path=True, report_type=report_type)
            path = path.replace(replace, "")

        path = isUnix(path)
        replace = isUnix(replace)

        if create_if_not_exists:
            path_exists(path)

        if complete:
            return path+replace
        
        return path, replace


def get_config_path():
    """Localiza o arquivo com as credenciais de acesso informado pela equipe da API

    Returns:
        - str: path completo do arquivo de configuração json
    """
    # path = os.getcwd()
    # path += r'\configs\rede-credentials.json'
    return r"https://firebasestorage.googleapis.com/v0/b/rede-api.appspot.com/o/server%2Frede-credentials.json?alt=media&token=09f1992d-3aaf-41da-ad70-262c07481d2b"

def get_report_path():
    return isUnix(os.getcwd()+r'\relatorio\reports')

def save_json(path, object):
    """Salva o json em local especificado por parâmetro

    Args:
       - path (str): caminho onde salva o arquivo
       - object (json): json com os dados
    """
    with open(path, 'w') as file:
        json.dump(object, file)

def send_json(path, object):
    requests.put(path, object)
    logger.info(f"enviado para {path}")
    logger.info(f"com sucesso")

def loads_json(object:dict):
    """transforma um dicionário em json e o retorna

    Args:
        - object (dict): dicionário com os dados

    Returns:
        - json: json
    """
    return json.loads(object)

def load_web_json(url):
    response = requests.get(url)
    return response.json()

def load_json(path:str):
    """lê e retorna um json pelo path especificado por parâmetro

    Args:
        - path (str): caminho para o json

    Returns:
        - json: json
    """
    with open(path, 'r') as file:
        data = json.load(file)
    return data
    
def is_future_date(date:dt):
    """verifica se é uma data futura, o que geraria um erro na requisiçao,
    ja que o arquivo seria inexistente

    Args:
        - date (datetime): data em formato datetime

    Raises:
        - ResourceWarning: Mensagem de erro
    """
    if date.month > dt.now().month:
        raise ResourceWarning("Data futura. impossível acessar arquivo")

def format_export(path:str) -> str:
    """
    Faz a validação e mudança para o diretório correto do arquivo excel exportado

    Returns: 
       - path completo para exportar .xlsx
    """
    path = path_exists(path)
    path = path.replace("connection", "vendas")
    return path

def find_desktop_path():
    """Busca o caminho da área de trabalho do desktop independente do s.o

    Returns:
        str: path com o caminho para o desktop
    """
    if os.name == "posix":
       return os.path.join(os.environ['HOME'], 'Área de Trabalho')
    return os.path.join(os.environ['USERPROFILE'], 'Desktop')   

def save_excel(df:pd.DataFrame, path:str, desktop:str):
        while True:
            try:
                estilo = {'font': Font(name='Arial', size=14, bold=False, color='FFFFFF'),
                            'fill': PatternFill(fill_type='solid', start_color='0FCF5F', end_color='FFFFFF'),
                            'border': Border(left=Side(border_style='thin', color='000000'),
                            right=Side(border_style='thin', color='000000'),
                            top=Side(border_style='thin', color='000000'),
                            bottom=Side(border_style='thin', color='000000')),
                            'alignment': Alignment(horizontal='center', vertical='center')}

                writer = pd.ExcelWriter(path, engine='openpyxl')
                df.to_excel(writer, index=False)
                workbook = writer.book
                worksheet = writer.sheets['Sheet1']

                for col_num, value in enumerate(df.columns.values):
                    worksheet.cell(1, col_num+1).value = value
                    for key in estilo:
                        setattr(worksheet.cell(1, col_num+1), key, estilo[key])
                workbook.save(path)

                # df.to_excel(path, index=False)
                shutil.copy(path, desktop)
                logger.info("Succesfully saved and copy to desktop")
                break
            except PermissionError:
                logger.warning("Erro de Permissão de escrita")
                logger.warning("Por favor feche o arquivo aberto ou veja com o adm do sistema")
                logger.warning(f"Arquivo: {path}")
                time.sleep(5)

def init_liberation():
    """
    Verifica se o arquivo com os dados do cliente informado existe.
    Caso contrário, encerra o programa
    """
    try:
        path = getPath()
        load_json(path)
    except ImportError:
        logger.error("Arquivo com companyNumber e infos inexistente")
        time.sleep(5)
        sys.exit()
    
    